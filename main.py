import os, time
import schedule
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

MESSENGER_URL = "https://www.messenger.com"
LOGIN_EMAIL = os.getenv("LOGIN_EMAIL")
LOGIN_PASSWORD = os.getenv("LOGIN_PASSWORD")


def setup_webdriver():
    print("Setting up web driver...")
    options = Options()
    options.add_argument("--headless=new")
    options.add_experimental_option(
        "prefs", {"profile.default_content_setting_values.notifications": 2}
    )
    browser = webdriver.Chrome(options=options)
    return browser


def login(browser):
    print("Waiting for login page to load...")
    WebDriverWait(browser, 10).until(
        EC.visibility_of_element_located((By.NAME, "email"))
    )
    email_input = browser.find_element(By.NAME, "email")
    password_input = browser.find_element(By.NAME, "pass")
    email_input.send_keys(LOGIN_EMAIL)
    password_input.send_keys(LOGIN_PASSWORD)
    login_button = browser.find_element(By.NAME, "login")
    login_button.click()
    print("Logging in...")


def close_popups(browser):
    print("Closing popups...")
    WebDriverWait(browser, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "div[aria-label='Close']"))
    )
    sync_history_popup_close_button = browser.find_element(
        By.CSS_SELECTOR, "div[aria-label='Close']"
    )
    sync_history_popup_close_button.click()

    WebDriverWait(browser, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'div[aria-label="Don\'t sync"]'))
    )
    continue_without_syncing_button = browser.find_element(
        By.CSS_SELECTOR, """div[aria-label="Don't sync"]"""
    )
    continue_without_syncing_button.click()
    print("Popups closed...")


def search_for_friend(browser, friend_name):
    print()
    print(f"Searching for {friend_name}...")
    WebDriverWait(browser, 10).until(
        EC.visibility_of_element_located(
            (By.CSS_SELECTOR, "input[aria-label='Search Messenger']")
        )
    )
    search_input = browser.find_element(
        By.CSS_SELECTOR, "input[aria-label='Search Messenger']"
    )
    search_input.send_keys(friend_name)
    WebDriverWait(browser, 10).until(
        EC.visibility_of_any_elements_located((By.CSS_SELECTOR, "li[role='option']"))
    )
    friends_options = browser.find_elements(By.CSS_SELECTOR, "li[role='option']")
    friend_option = [
        option for option in friends_options if option.text == friend_name
    ][0]
    friend_option.click()


def send_message(browser):
    print("Sending Message...")
    WebDriverWait(browser, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "div[aria-label='Message']"))
    )
    message_box = browser.find_element(By.CSS_SELECTOR, "div[aria-label='Message']")
    message_box.send_keys(f"Happy Birthday. Many Many Happy Returns of the day.")
    send_button = browser.find_element(
        By.CSS_SELECTOR, "div[aria-label='Press enter to send']"
    )
    send_button.click()


def get_current_day():
    current_day = datetime.today().strftime("%Y-%m-%d").split("-")
    current_day = datetime(
        int(current_day[0]), int(current_day[1]), int(current_day[2])
    )
    return current_day


def get_people_with_birthdays():
    wb = load_workbook("birthdays.xlsx")
    sheet = wb.active
    name_birthdays = [row for row in sheet.iter_rows(max_col=2, values_only=True)]
    wb.close()
    current_day = get_current_day()
    people_with_birthday_today = []
    for name_birthday in name_birthdays:
        name_birthday_temp = str(datetime.today().year) + "-" + name_birthday[1]
        name_birthday_temp = datetime.strptime(name_birthday_temp, format="%Y-%m-%d")
        if current_day == name_birthday_temp:
            people_with_birthday_today.append(name_birthday[0])
    return people_with_birthday_today


def wish_happy_birthday():
    people_with_birthday_today = get_people_with_birthdays()
    print(people_with_birthday_today)
    if not people_with_birthday_today:
        print("Noone has a birthday today!")
    else:
        for name in people_with_birthday_today:
            print(f"\t{name} has birthday today.")
        print()
        browser = setup_webdriver()
        print("Loading 'messenger.com'...")
        browser.get(MESSENGER_URL)
        login(browser)
        close_popups(browser)
        for name in people_with_birthday_today:
            search_for_friend(browser, friend_name=name)
            send_message(browser)
            time.sleep(5)
            print("Message Sent...")
            browser.close()


def main():
    schedule.every().day.at("15:14").do(wish_happy_birthday)
    print("Program Running...\n\n")
    while True:
        schedule.run_pending()
        time.sleep(30)


if __name__ == "__main__":
    main()
